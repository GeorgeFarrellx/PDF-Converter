# Version: 2.29
import os
import glob
import re
import importlib.util
import traceback
from datetime import datetime, timedelta, date
import sys
import hashlib
import json
import platform
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

_PDFPLUMBER_CACHE = None
_PDFPLUMBER_ERROR_SHOWN = False


# ----------------------------
# CONFIG (edit these as needed)
# ----------------------------

DEFAULT_OUTPUT_FOLDER = ""
PARSERS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Parsers")
LOGS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logs")
BANK_OPTIONS = [
    "Select bank...",
    "Nationwide",
    "RBS",
    "Barclays",
    "Halifax",
    "Lloyds",
    "Starling",
    "TSB",
    "NatWest",
    "Monzo",
    "Santander",
    "HSBC",
    "Zempler Bank",
]


# ----------------------------
# Utilities
# ----------------------------

def ensure_folder(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _show_dependency_error(message: str) -> None:
    try:
        messagebox.showerror("Missing Dependency", message)
    except Exception:
        print(message, file=sys.stderr)


def _require_pdfplumber(show_error: bool = True):
    global _PDFPLUMBER_CACHE, _PDFPLUMBER_ERROR_SHOWN
    if _PDFPLUMBER_CACHE is not None:
        return _PDFPLUMBER_CACHE
    try:
        import pdfplumber
    except Exception as e:
        if show_error and not _PDFPLUMBER_ERROR_SHOWN:
            _show_dependency_error(
                "pdfplumber is required for PDF text extraction.\n\n"
                "Install it with:\n"
                "  python -m pip install pandas openpyxl pdfplumber\n\n"
                f"Original error: {e}"
            )
            _PDFPLUMBER_ERROR_SHOWN = True
        return None
    _PDFPLUMBER_CACHE = pdfplumber
    return pdfplumber


def make_unique_path(path: str) -> str:
    """If path exists, append ' (2)', ' (3)'... before extension."""
    if not os.path.exists(path):
        return path

    base, ext = os.path.splitext(path)
    n = 2
    while True:
        candidate = f"{base} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def normalize_bank_name_for_module(bank: str) -> str:
    normalized = bank.strip().lower()
    if normalized == "zempler bank":
        return "zempler"
    return normalized


def load_parser_module(bank: str):
    bank_module_name = normalize_bank_name_for_module(bank)

    parser_path = os.path.join(PARSERS_DIR, f"{bank_module_name}.py")

    if not os.path.exists(parser_path):
        pattern = os.path.join(PARSERS_DIR, f"{bank_module_name}-*.py")
        matches = sorted(glob.glob(pattern))
        if matches:
            parser_path = matches[-1]

    if not os.path.exists(parser_path):
        raise FileNotFoundError(
    f"No parser found for bank '{bank}'. Expected either:\n"
    f"  - {os.path.join(PARSERS_DIR, bank_module_name + '.py')}\n"
    f"  - {os.path.join(PARSERS_DIR, bank_module_name + '-<version>.py')}\n\n"
    f"Tried: {parser_path}"
)

    module_key = os.path.splitext(os.path.basename(parser_path))[0]

    spec = importlib.util.spec_from_file_location(f"parsers.{module_key}", parser_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load parser module from {parser_path}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    if not hasattr(module, "extract_transactions"):
        raise AttributeError(
            f"Parser '{parser_path}' does not define extract_transactions(pdf_path)."
        )

    return module


def auto_detect_bank_from_pdf(pdf_path: str) -> str | None:
    """Best-effort bank detection.

    To handle cover/summary sheets, scan the first few pages (instead of only page 1).
    Keep this light-weight: extract_text() only, simple substring checks.

    NOTE: Avoid false positives from transaction descriptions that mention other banks.
    We therefore only scan the header portion of each page (stop at the transaction table header).

    Some PDFs (e.g. NatWest transaction exports) only include bank identifiers in the footer of the last page,
    so we also scan a small footer snippet from the final page.
    """
    pdfplumber = _require_pdfplumber()
    if pdfplumber is None:
        return None

    MAX_PAGES = 4
    MAX_LINES_PER_PAGE = 60
    LAST_PAGE_FOOTER_LINES = 30
    header_re = re.compile(r"date.*(description|details).*balance", re.IGNORECASE)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None

            texts: list[str] = []
            for p in pdf.pages[: min(len(pdf.pages), MAX_PAGES)]:
                txt = ""
                try:
                    txt = p.extract_text() or ""
                except Exception:
                    txt = ""

                if not txt:
                    try:
                        words = p.extract_words() or []
                        if words:
                            txt = " ".join(w.get("text", "") for w in words if w.get("text"))
                    except Exception:
                        pass

                if not txt:
                    try:
                        chars = getattr(p, "chars", None) or []
                        if chars:
                            txt = "".join(c.get("text", "") for c in chars if c.get("text"))
                    except Exception:
                        pass

                if txt:
                    lines = txt.splitlines()
                    cut = min(len(lines), MAX_LINES_PER_PAGE)
                    for j, line in enumerate(lines[:cut]):
                        if header_re.search(line or ""):
                            cut = j + 1
                            break
                    txt = (chr(10)).join(lines[:cut]) if lines else txt
                    texts.append(txt)

            t = (chr(10)).join(texts).lower()

            last_footer = ""
            try:
                last_page = pdf.pages[-1]
                last_txt = ""
                try:
                    last_txt = last_page.extract_text() or ""
                except Exception:
                    last_txt = ""

                if last_txt:
                    last_lines = last_txt.splitlines()
                    if last_lines:
                        last_footer = (chr(10)).join(last_lines[-LAST_PAGE_FOOTER_LINES:])
            except Exception:
                last_footer = ""

        footer_low = (last_footer or "").lower()
        natwest_probe = t + (chr(10) + footer_low if footer_low else "")
        zempler_probe = t + (chr(10) + footer_low if footer_low else "")

        if "hsbc" in t or "hsbc uk" in t or "hsbc bank" in t:
            return "HSBC"

        if "halifax" in t:
            return "Halifax"

        if "lloyds" in t or "lloyds bank" in t:
            return "Lloyds"

        if "starling" in t or "starling bank" in t:
            return "Starling"

        if (
            "tsb bank" in t
            or "tsb bank plc" in t
            or "tsb.co.uk" in t
            or "www.tsb.co.uk" in t
            or "tsb internet banking" in t
        ):
            return "TSB"

        if (
            "national westminster bank" in natwest_probe
            or "natwest online transactions service" in natwest_probe
            or "downloaded from the natwest online transactions service" in natwest_probe
            or "natwest.com" in natwest_probe
            or "natwest" in t
            or "nat west" in t
        ):
            return "NatWest"

        if (
            "santander" in t
            or "santander uk" in t
            or "santander uk plc" in t
            or "santander.co.uk" in t
            or "www.santander.co.uk" in t
            or "abbygb2l" in t
            or "abby gb2l" in t
        ):
            return "Santander"

        if (
            "monzo" in t
            or "monzgb2l" in t
            or "monzo bank limited" in t
            or "monzo.com" in t
            or "04-00-04" in t
            or "040004" in t
        ):
            return "Monzo"

        if "nationwide" in t or "flexbasic" in t or "your flexbasic account" in t:
            return "Nationwide"

        if "royal bank of scotland" in t or "rbs" in t or "rbs.co.uk" in t:
            return "RBS"
        if "period covered" in t and "new balance" in t and "previous balance" in t:
            return "RBS"

        if (
            "barclays" in t
            or "barclays bank" in t
            or "barclays.co.uk" in t
            or "bukbgb22" in t
            or ("iban" in t and "bukb" in t)
            or "your business current account" in t
            or (
                "date" in t
                and "description" in t
                and "money out" in t
                and "money in" in t
                and "balance" in t
            )
        ):
            return "Barclays"

        if "zempler" in zempler_probe:
            return "Zempler Bank"

    except Exception:
        return None

    return None


def extract_barclays_statement_period(pdf_path: str) -> tuple[date | None, date | None]:
    pdfplumber = _require_pdfplumber()
    if pdfplumber is None:
        return (None, None)

    MAX_PAGES = 4

    month_map = {
        "jan": 1,
        "feb": 2,
        "mar": 3,
        "apr": 4,
        "may": 5,
        "jun": 6,
        "jul": 7,
        "aug": 8,
        "sep": 9,
        "oct": 10,
        "nov": 11,
        "dec": 12,
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return (None, None)

            texts: list[str] = []
            for p in pdf.pages[: min(len(pdf.pages), MAX_PAGES)]:
                try:
                    texts.append(p.extract_text() or "")
                except Exception:
                    texts.append("")
        full = (chr(10)).join(texts)
        low = full.lower()

        idx = low.find("at a glance")
        region = full[idx : idx + 1200] if idx != -1 else full

        m = re.search(
            r"(\d{1,2})\s+([A-Za-z]{3})\s*-\s*(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})",
            region,
        )
        if m:
            sd = int(m.group(1))
            sm = month_map.get(m.group(2).lower())
            ed = int(m.group(3))
            em = month_map.get(m.group(4).lower())
            yr = int(m.group(5))
            if sm and em:
                start = date(yr, sm, sd)
                end_year = yr + 1 if em < sm else yr
                end = date(end_year, em, ed)
                return (start, end)

        m = re.search(
            r"(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s*-\s*(\d{1,2})\s+([A-Za-z]{3})\b",
            region,
        )
        if m:
            sd = int(m.group(1))
            sm = month_map.get(m.group(2).lower())
            sy = int(m.group(3))
            ed = int(m.group(4))
            em = month_map.get(m.group(5).lower())

            if sm and em:
                tail = region[m.end() : m.end() + 200]
                my = re.search(r"\b(20\d{2})\b", tail)
                end_year = int(my.group(1)) if my else (sy + 1 if em < sm else sy)

                start = date(sy, sm, sd)
                end = date(end_year, em, ed)
                return (start, end)

    except Exception:
        return (None, None)

    return (None, None)


def extract_statement_period_from_pdf(pdf_path: str, bank_hint: str = "") -> tuple[date | None, date | None]:
    """Extract statement period dates for continuity/chain logic.

    Starling format supported (page 1):
      "Summary dd/mm/yyyy - dd/mm/yyyy"

    Returns (period_start, period_end) as datetime.date, or (None, None).
    """
    pdfplumber = _require_pdfplumber()
    if pdfplumber is None:
        return (None, None)

    hint = (bank_hint or "").lower()
    try_starling = ("starling" in hint) or (hint.strip() == "")
    if not try_starling:
        return (None, None)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return (None, None)
            try:
                txt = pdf.pages[0].extract_text() or ""
            except Exception:
                txt = ""

        if not txt:
            return (None, None)

        m = re.search(r"Summary\s+(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})", txt, re.IGNORECASE)
        if not m:
            return (None, None)

        ps = datetime.strptime(m.group(1), "%d/%m/%Y").date()
        pe = datetime.strptime(m.group(2), "%d/%m/%Y").date()
        return (ps, pe)

    except Exception:
        return (None, None)



def parse_dnd_event_files(data: str) -> list[str]:
    files = []
    token = ""
    in_brace = False

    for ch in data:
        if ch == "{":
            in_brace = True
            token = ""
        elif ch == "}":
            in_brace = False
            if token:
                files.append(token)
                token = ""
        elif ch == " " and not in_brace:
            if token:
                files.append(token)
                token = ""
        else:
            token += ch

    if token:
        files.append(token)

    cleaned = []
    for f in files:
        f = f.strip().strip('"')
        if f:
            cleaned.append(f)
    return cleaned


def is_pdf(path: str) -> bool:
    return path.lower().endswith(".pdf")


def get_client_name_from_pdf(pdf_path: str) -> str:
    pdfplumber = _require_pdfplumber()
    if pdfplumber is None:
        return ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return ""
            text = pdf.pages[0].extract_text() or ""
    except Exception:
        return ""

    titles = ("Mr ", "Mrs ", "Ms ", "Miss ", "Dr ")
    for raw in (text.splitlines() if text else []):
        line = (raw or "").strip()
        if not line:
            continue
        if any(ch.isdigit() for ch in line):
            continue
        if line.startswith(titles):
            if "statement" in line.lower():
                continue
            return line

    return ""


def sanitize_filename(name: str) -> str:
    bad = r'<>:/\\|?*"'
    out = "".join("_" if ch in bad else ch for ch in (name or ""))
    out = " ".join(out.split())
    return out.strip().strip(".")


def build_output_filename(client_name: str, date_min, date_max) -> str:
    cname = (client_name or "").strip() or "Transactions"
    cname = sanitize_filename(cname.upper())

    period = ""
    try:
        if date_min and date_max:
            period = f"{date_min.strftime('%d.%m.%y')} - {date_max.strftime('%d.%m.%y')}"
    except Exception:
        period = ""

    if period:
        return f"{cname} {period}.xlsx"
    return f"{cname}.xlsx"


def compute_statement_fingerprint(transactions: list[dict]) -> str | None:
    if not transactions:
        return None

    def _norm_text(v):
        if v is None:
            return ""
        s = str(v)
        s = " ".join(s.split())
        return s.upper().strip()

    def _norm_date(v):
        if v is None or v == "":
            return ""
        try:
            if hasattr(v, "to_pydatetime"):
                v = v.to_pydatetime()
        except Exception:
            pass
        if hasattr(v, "strftime"):
            try:
                return v.strftime("%Y-%m-%d")
            except Exception:
                pass
        return str(v).strip()

    def _norm_money(v):
        if v is None or v == "":
            return ""
        try:
            return f"{round(float(v), 2):.2f}"
        except Exception:
            return str(v).strip()

    rows: list[str] = []
    for t in transactions:
        if not isinstance(t, dict):
            continue
        row = "|".join(
            [
                _norm_date(t.get("Date")),
                _norm_text(t.get("Transaction Type")),
                _norm_text(t.get("Description")),
                _norm_money(t.get("Amount")),
                _norm_money(t.get("Balance")),
            ]
        )
        rows.append(row)

    if not rows:
        return None

    rows.sort()
    payload = "\n".join(rows).encode("utf-8", errors="ignore")
    return hashlib.sha1(payload).hexdigest()


def find_duplicate_statements(recon_results: list[dict]) -> list[list[dict]]:
    fp_map: dict[str, list[dict]] = {}
    for r in recon_results or []:
        fp = r.get("fingerprint")
        if not fp:
            continue
        fp_map.setdefault(fp, []).append(r)

    return [grp for grp in fp_map.values() if len(grp) > 1]


# ----------------------------
# Excel output
# ----------------------------

def _find_rules_file(folder: str, base_name: str) -> str | None:
    if not folder:
        return None
    xlsx_path = os.path.join(folder, f"{base_name}.xlsx")
    if os.path.exists(xlsx_path):
        return xlsx_path
    csv_path = os.path.join(folder, f"{base_name}.csv")
    if os.path.exists(csv_path):
        return csv_path
    return None


def _read_rules_csv(path: str, pd):
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]
    last_err = None
    for enc in encodings:
        try:
            return pd.read_csv(path, encoding=enc, dtype=str, keep_default_na=False)
        except Exception as e:
            last_err = e
    print(f"[core] WARNING: Failed to read rules CSV '{path}': {last_err}")
    return None


def _excel_list_separator() -> str:
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Control Panel\\International") as k:
            val, _ = winreg.QueryValueEx(k, "sList")
            if isinstance(val, str) and val.strip():
                return val.strip()
    except Exception:
        pass
    return ","


def _load_rules(path: str, pd) -> list[dict]:
    if path.lower().endswith(".csv"):
        df = _read_rules_csv(path, pd)
        if df is None:
            return []
    else:
        try:
            excel = pd.ExcelFile(path)
            sheet_name = "Category Rules" if "Category Rules" in excel.sheet_names else excel.sheet_names[0]
            df = pd.read_excel(path, sheet_name=sheet_name)
        except Exception as e:
            print(f"[core] WARNING: Failed to read rules file '{path}': {e}")
            return []

    records: list[dict] = []

    def _as_bool(v, default: bool = True) -> bool:
        if v is None:
            return default
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in {"", "nan"}:
            return default
        if s in {"true", "1", "yes", "y", "on"}:
            return True
        if s in {"false", "0", "no", "n", "off"}:
            return False
        return default

    def _as_priority(v) -> int:
        if v is None:
            return 9999
        try:
            s = str(v).strip()
            if not s or s.lower() == "nan":
                return 9999
            return int(float(s))
        except Exception:
            return 9999

    for _, row in df.iterrows():
        priority = _as_priority(row.get("Priority") if "Priority" in row else None)
        category = str(row.get("Category") if "Category" in row else "").strip()
        match_type = str(row.get("Match Type") if "Match Type" in row else "").strip()
        pattern = str(row.get("Pattern") if "Pattern" in row else "").strip()
        direction = str(row.get("Direction") if "Direction" in row else "").strip()
        txn_type_contains = str(row.get("Txn Type Contains") if "Txn Type Contains" in row else "").strip()
        active = _as_bool(row.get("Active") if "Active" in row else None, default=True)

        if not active:
            continue
        if not category or category.lower() == "nan":
            continue
        if not pattern or pattern.lower() == "nan":
            continue

        records.append(
            {
                "Priority": priority,
                "Category": category,
                "Match Type": match_type,
                "Pattern": pattern,
                "Direction": direction,
                "Txn Type Contains": txn_type_contains,
            }
        )

    records.sort(key=lambda r: r.get("Priority", 9999))
    return records


def _rule_matches(txn: dict, rule: dict) -> bool:
    description = str(txn.get("Description", "") or "")
    txn_type = str(txn.get("Transaction Type", "") or "")
    pattern = str(rule.get("Pattern", "") or "")
    if not pattern:
        return False

    txn_type_contains = str(rule.get("Txn Type Contains", "") or "").strip()
    if txn_type_contains and txn_type_contains.lower() != "nan":
        if txn_type_contains.lower() not in txn_type.lower():
            return False

    direction = str(rule.get("Direction", "") or "").strip().upper()
    if direction and direction != "ANY":
        try:
            amount = float(txn.get("Amount", 0) or 0)
        except Exception:
            amount = 0.0
        if direction == "DEBIT" and amount >= 0:
            return False
        if direction == "CREDIT" and amount <= 0:
            return False

    match_type = str(rule.get("Match Type", "") or "").strip().lower() or "contains"
    desc_cmp = description.strip()
    patt_cmp = pattern.strip()
    desc_low = desc_cmp.lower()
    patt_low = patt_cmp.lower()

    if match_type == "exact":
        return desc_low == patt_low
    if match_type == "startswith":
        return desc_low.startswith(patt_low)
    if match_type == "endswith":
        return desc_low.endswith(patt_low)
    if match_type == "regex":
        try:
            return re.search(patt_cmp, desc_cmp, re.IGNORECASE) is not None
        except Exception:
            return False
    return patt_low in desc_low


def _apply_global_categorisation(transactions: list[dict], pd) -> None:
    global_folder = os.path.dirname(os.path.abspath(__file__))
    global_rules_path = _find_rules_file(global_folder, "Global Categorisation Rules")
    if not global_rules_path:
        missing_csv = os.path.join(global_folder, "Global Categorisation Rules.csv")
        print(f"[core] WARNING: Global categorisation rules file not found: '{missing_csv}'")
    rules = _load_rules(global_rules_path, pd) if global_rules_path else []

    for txn in transactions:
        if not isinstance(txn, dict):
            continue
        if "Global Category" not in txn:
            txn["Global Category"] = ""
        if str(txn.get("Global Category", "") or "").strip():
            continue
        for rule in rules:
            if _rule_matches(txn, rule):
                txn["Global Category"] = rule.get("Category", "")
                break


def _sha256_file(path: str) -> str | None:
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            while True:
                chunk = f.read(65536)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None


def _write_log_text(prefix: str, content: str) -> str | None:
    try:
        ensure_folder(LOGS_DIR)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(LOGS_DIR, f"{prefix}_{ts}.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return path
    except Exception:
        return None


def _write_log_json(prefix: str, obj) -> str | None:
    try:
        ensure_folder(LOGS_DIR)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(LOGS_DIR, f"{prefix}_{ts}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, indent=2, ensure_ascii=False)
        return path
    except Exception:
        return None


def _write_categorisation_evidence_instructions(audit: dict, audit_json_path: str | None) -> str | None:
    try:
        ensure_folder(LOGS_DIR)
        ts = audit.get("audit_ts") or datetime.now().strftime("%Y%m%d_%H%M%S")
        instructions_path = os.path.join(LOGS_DIR, f"categorisation_evidence_pack_instructions_{ts}.txt")

        version_text = None
        version_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VERSION.txt")
        try:
            with open(version_path, "r", encoding="utf-8") as vf:
                version_text = vf.read().strip()
        except Exception:
            version_text = None

        lines = []
        lines.append("Categorisation Evidence Pack Instructions")
        lines.append("========================================")
        lines.append("")
        lines.append(f"Output workbook path: {audit.get('output_path')}")
        if version_text:
            lines.append(f"App version: {version_text}")
        else:
            lines.append(f"core.py path: {audit.get('core_file')}")
            lines.append(f"core.py sha256: {audit.get('core_sha256')}")
        lines.append(f"Python version: {audit.get('python_version')}")
        lines.append(f"pandas version: {audit.get('pandas_version')}")
        lines.append(f"openpyxl version: {audit.get('openpyxl_version')}")
        lines.append(f"Platform: {audit.get('platform')}")
        lines.append(f"excel_list_separator_detected: {audit.get('excel_list_separator_detected')}")
        lines.append("")
        lines.append("If Excel shows 'needs repairs'")
        lines.append("------------------------------")
        lines.append("1) Open the workbook in Excel.")
        lines.append("2) Click Repair.")
        lines.append("3) Click Show Repairs.")
        lines.append("4) Copy/paste the full 'Repaired Records' text into your report.")
        lines.append("")
        lines.append("Send these files")
        lines.append("---------------")
        lines.append(f"- Output workbook: {audit.get('output_path')}")
        lines.append(f"- Categorisation audit JSON: {audit_json_path}")
        artifact_paths = audit.get("artifact_paths") or {}
        for _, artifact_path in artifact_paths.items():
            lines.append(f"- Audit artifact: {artifact_path}")
        lines.append("- Include the latest recon log from Logs/ (recon_*.txt) if present.")
        lines.append("- Include any crash logs from Logs/ (crash_*.txt) if present.")
        lines.append("- Include the Global Categorisation Rules.csv file used for the run.")

        with open(instructions_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return instructions_path
    except Exception:
        return None


def _audit_xlsx_categorisation(output_path: str) -> dict:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    abs_output = os.path.abspath(output_path)
    ensure_folder(LOGS_DIR)
    audit = {
        "audit_ts": ts,
        "output_path": abs_output,
        "enable_categorisation": True,
        "core_file": os.path.abspath(__file__),
        "core_sha256": _sha256_file(os.path.abspath(__file__)),
        "excel_list_separator_detected": _excel_list_separator(),
        "python_version": sys.version,
        "platform": platform.platform(),
        "pandas_version": None,
        "openpyxl_version": None,
        "artifact_paths": {},
        "transaction_data_sheet_xml": None,
        "categorisation_rules_sheet_xml": None,
        "transaction_data_formula_count": 0,
        "transaction_data_formula_samples": [],
        "categorisation_rules_formula_count": 0,
        "tables": [],
    }

    try:
        import pandas as _pd

        audit["pandas_version"] = getattr(_pd, "__version__", None)
    except Exception:
        pass

    try:
        import openpyxl as _openpyxl

        audit["openpyxl_version"] = getattr(_openpyxl, "__version__", None)
    except Exception:
        pass

    workbook_xml = None
    workbook_rels_xml = None
    transaction_sheet_xml_raw = None
    tables_summary_lines = []

    with zipfile.ZipFile(abs_output, "r") as zf:
        names = set(zf.namelist())
        rel_target_map = {}
        sheet_targets = {}

        if "xl/workbook.xml" in names:
            workbook_xml = zf.read("xl/workbook.xml").decode("utf-8", errors="replace")
        if "xl/_rels/workbook.xml.rels" in names:
            workbook_rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")

        if workbook_rels_xml:
            rel_root = ET.fromstring(workbook_rels_xml)
            for rel in rel_root.findall("rel:Relationship", ns):
                rel_id = rel.attrib.get("Id")
                target = rel.attrib.get("Target", "")
                if rel_id and target:
                    if target.startswith("/"):
                        target = target.lstrip("/")
                    if not target.startswith("xl/"):
                        target = f"xl/{target}"
                    rel_target_map[rel_id] = target

        if workbook_xml:
            wb_root = ET.fromstring(workbook_xml)
            sheets = wb_root.find("main:sheets", ns)
            if sheets is not None:
                for sheet in sheets.findall("main:sheet", ns):
                    name = sheet.attrib.get("name")
                    rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                    if name and rid and rid in rel_target_map:
                        sheet_targets[name] = rel_target_map[rid]

        transaction_sheet_xml = sheet_targets.get("Transaction Data")
        categorisation_rules_sheet_xml = sheet_targets.get("Client Categorisation Rules")
        audit["transaction_data_sheet_xml"] = transaction_sheet_xml
        audit["categorisation_rules_sheet_xml"] = categorisation_rules_sheet_xml

        def _read_formula_info(sheet_path: str):
            if not sheet_path or sheet_path not in names:
                return (0, [])
            xml_text = zf.read(sheet_path).decode("utf-8", errors="replace")
            root = ET.fromstring(xml_text)
            count = 0
            samples = []
            for cell_node in root.findall(".//main:c", ns):
                f_node = cell_node.find("main:f", ns)
                if f_node is None:
                    continue
                count += 1
                if len(samples) < 20:
                    samples.append({
                        "ref": cell_node.attrib.get("r"),
                        "formula": f_node.text or "",
                    })
            return (count, samples)

        tx_count, tx_samples = _read_formula_info(transaction_sheet_xml)
        rules_count, _ = _read_formula_info(categorisation_rules_sheet_xml)
        audit["transaction_data_formula_count"] = tx_count
        audit["transaction_data_formula_samples"] = tx_samples
        audit["categorisation_rules_formula_count"] = rules_count

        table_paths = sorted(n for n in names if n.startswith("xl/tables/table") and n.endswith(".xml"))
        for table_path in table_paths:
            table_xml = zf.read(table_path).decode("utf-8", errors="replace")
            t_root = ET.fromstring(table_xml)
            display_name = t_root.attrib.get("displayName")
            ref = t_root.attrib.get("ref")
            columns = []
            calculated_column_formulas = {}
            calculated_column_formula_attrs = {}
            table_columns = t_root.find("main:tableColumns", ns)
            if table_columns is not None:
                for col in table_columns.findall("main:tableColumn", ns):
                    column_name = col.attrib.get("name")
                    columns.append(column_name)
                    calculated_formula_node = col.find("main:calculatedColumnFormula", ns)
                    if calculated_formula_node is not None and calculated_formula_node.text is not None:
                        calculated_column_formulas[column_name] = calculated_formula_node.text
                        calculated_column_formula_attrs[column_name] = dict(calculated_formula_node.attrib)
            audit["tables"].append(
                {
                    "path": table_path,
                    "displayName": display_name,
                    "ref": ref,
                    "tableColumns": columns,
                    "calculatedColumnFormulas": calculated_column_formulas,
                    "calculatedColumnFormulaAttrs": calculated_column_formula_attrs,
                }
            )
            tables_summary_lines.append(f"{table_path} | displayName={display_name} | ref={ref}")
            tables_summary_lines.append(", ".join(str(c) for c in columns))
            tables_summary_lines.append("")

        if transaction_sheet_xml and transaction_sheet_xml in names:
            transaction_sheet_xml_raw = zf.read(transaction_sheet_xml).decode("utf-8", errors="replace")

    if transaction_sheet_xml_raw is not None:
        artifact_path = os.path.abspath(os.path.join(LOGS_DIR, f"categorisation_audit_{ts}_transaction_sheet.xml"))
        with open(artifact_path, "w", encoding="utf-8") as f:
            f.write(transaction_sheet_xml_raw)
        audit["artifact_paths"]["transaction_sheet_xml"] = artifact_path
    if workbook_xml is not None:
        artifact_path = os.path.abspath(os.path.join(LOGS_DIR, f"categorisation_audit_{ts}_workbook.xml"))
        with open(artifact_path, "w", encoding="utf-8") as f:
            f.write(workbook_xml)
        audit["artifact_paths"]["workbook_xml"] = artifact_path
    if workbook_rels_xml is not None:
        artifact_path = os.path.abspath(os.path.join(LOGS_DIR, f"categorisation_audit_{ts}_workbook.rels"))
        with open(artifact_path, "w", encoding="utf-8") as f:
            f.write(workbook_rels_xml)
        audit["artifact_paths"]["workbook_rels"] = artifact_path
    if tables_summary_lines:
        artifact_path = os.path.abspath(os.path.join(LOGS_DIR, f"categorisation_audit_{ts}_tables_summary.txt"))
        with open(artifact_path, "w", encoding="utf-8") as f:
            f.write("\n".join(tables_summary_lines))
        audit["artifact_paths"]["tables_summary"] = artifact_path

    return audit

def save_transactions_to_excel(transactions: list[dict], output_path: str, client_name: str = "", header_period_start=None, header_period_end=None, enable_categorisation: bool = True):
    if not transactions:
        raise ValueError("No transactions found!")

    try:
        import pandas as pd
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
        from openpyxl.styles import Border, Side
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.worksheet.filters import AutoFilter
    except Exception as e:
        _show_dependency_error(
            "pandas and openpyxl are required for Excel output.\n\n"
            "Install them with:\n"
            "  python -m pip install pandas openpyxl pdfplumber\n\n"
            f"Original error: {e}"
        )
        return None

    if enable_categorisation:
        try:
            _apply_global_categorisation(transactions, pd)
        except Exception:
            pass

    df = pd.DataFrame(transactions)
    if "T/N" not in df.columns:
        df.insert(0, "T/N", range(1, len(df) + 1))

    required = ["Date", "Transaction Type", "Description", "Amount", "Balance"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Parser output missing columns: {missing}")

    if "Global Category" not in df.columns:
        df["Global Category"] = None
    if "Manual Category" not in df.columns:
        df["Manual Category"] = ""
    if "Client Specific Category" not in df.columns:
        df["Client Specific Category"] = None
    if "Final Category" not in df.columns:
        df["Final Category"] = None

    df = df[["T/N", "Date", "Transaction Type", "Description", "Amount", "Balance", "Global Category", "Client Specific Category", "Manual Category", "Final Category"]]

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce")

    if enable_categorisation:
        df["Global Category"] = df["Global Category"].where(df["Global Category"].notna(), None)
    else:
        df["Global Category"] = ""

    df["Client Specific Category"] = None
    df["Final Category"] = None

    ensure_folder(os.path.dirname(output_path))
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        wb = writer.book
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True

        borderless_side = Side(style=None)
        borderless_border = Border(
            left=borderless_side,
            right=borderless_side,
            top=borderless_side,
            bottom=borderless_side,
        )
        borderless_dxf = DifferentialStyle(border=borderless_border)
        borderless_dxf_id = wb._differential_styles.add(borderless_dxf)

        df.to_excel(writer, index=False, sheet_name="Transaction Data")

        ws = writer.sheets["Transaction Data"]

        def _coerce_header_period_date(v):
            try:
                if v is None or v == "":
                    return None
                if hasattr(v, "to_pydatetime"):
                    v = v.to_pydatetime()
                if isinstance(v, datetime):
                    return v.date()
                if isinstance(v, date):
                    return v
            except Exception:
                return None
            return None

        hp_start = _coerce_header_period_date(header_period_start)
        hp_end = _coerce_header_period_date(header_period_end)

        use_start = hp_start if hp_start else df["Date"].min()
        use_end = hp_end if hp_end else df["Date"].max()

        period = ""
        try:
            if pd.notna(use_start) and pd.notna(use_end):
                period = f"{use_start.strftime('%d/%m/%y')} - {use_end.strftime('%d/%m/%y')}"
        except Exception:
            period = ""

        left_text = (client_name or "").strip()
        center_text = "Transaction Data"
        right_text = period

        for hdr in (ws.oddHeader, ws.evenHeader, ws.firstHeader):
            hdr.left.text = left_text
            hdr.center.text = center_text
            hdr.right.text = right_text

        last_row = ws.max_row
        last_col = ws.max_column
        if last_row >= 2 and last_col >= 1:
            ref = f"A1:{get_column_letter(last_col)}{last_row}"
            table = Table(displayName="TransactionData", ref=ref)
            table.tableColumns = [
                TableColumn(id=idx, name=header)
                for idx, header in enumerate(df.columns, start=1)
            ]

            style = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            table.totalsRowShown = False
            table.tableStyleInfo = style
            table.autoFilter = AutoFilter(ref=table.ref)
            table.headerRowBorderDxfId = borderless_dxf_id
            table.tableBorderDxfId = borderless_dxf_id
            table.totalsRowBorderDxfId = borderless_dxf_id
            ws.add_table(table)

            no_border = Border()
            for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
                for cell in row:
                    cell.border = no_border

        ws_rules = wb.create_sheet("Client Categorisation Rules")
        ws_rules.append(["Priority", "Category", "Pattern", "Direction", "Txn Type Contains", "Active", "Notes"])
        ws_rules.append([10, "Tools & Materials", "ELECTRICAL", "ANY", "", True, "TEMP seed rule for testing"])
        ws_rules.append([11, "Tools & Materials", "SCREWFIX", "ANY", "", True, "TEMP seed rule for testing"])
        ws_rules.append([1000, "", "", "ANY", "", False, "Add client rules here. Pattern is 'contains' (case-insensitive)."])
        rules_table = Table(displayName="ClientRules", ref="A1:G4")
        rules_style = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        rules_table.totalsRowShown = False
        rules_table.tableStyleInfo = rules_style
        rules_table.headerRowBorderDxfId = borderless_dxf_id
        rules_table.tableBorderDxfId = borderless_dxf_id
        rules_table.totalsRowBorderDxfId = borderless_dxf_id
        ws_rules.add_table(rules_table)
        ws_rules.freeze_panes = "A2"

        gbp_accounting = "_-£* #,##0.00_-;[Red]-£* #,##0.00_-;_-£* \"-\"??_-;_-@_-"

        header_to_col = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            if isinstance(header_val, str) and header_val.strip():
                header_to_col[header_val.strip()] = col_idx

        tn_col = header_to_col.get("T/N")
        date_col = header_to_col.get("Date")
        amt_col = header_to_col.get("Amount")
        bal_col = header_to_col.get("Balance")
        desc_col = header_to_col.get("Description")
        txn_type_col = header_to_col.get("Transaction Type")
        global_cat_col = header_to_col.get("Global Category")
        client_specific_col = header_to_col.get("Client Specific Category")
        manual_col = header_to_col.get("Manual Category")
        final_col = header_to_col.get("Final Category")

        disable_client_specific_formula_for_diagnostics = False
        sep = _excel_list_separator()

        max_r = ws.max_row
        if date_col:
            for r in range(2, max_r + 1):
                ws.cell(row=r, column=date_col).number_format = "dd/mm/yyyy"
        if amt_col:
            for r in range(2, max_r + 1):
                ws.cell(row=r, column=amt_col).number_format = gbp_accounting
        if bal_col:
            for r in range(2, max_r + 1):
                ws.cell(row=r, column=bal_col).number_format = gbp_accounting
        if (
            client_specific_col
            and desc_col
            and not disable_client_specific_formula_for_diagnostics
        ):
            desc_letter = get_column_letter(desc_col)
            for r in range(2, max_r + 1):
                desc_ref = f"{desc_letter}{r}"
                cond_expr = (
                    f"((ClientRules[[#Data],[Active]]=TRUE)+(ClientRules[[#Data],[Active]]=\"\"))"
                    f"*(ClientRules[[#Data],[Pattern]]<>\"\")"
                    f"*(ClientRules[[#Data],[Priority]]<>\"\")"
                    f"*ISNUMBER(SEARCH(ClientRules[[#Data],[Pattern]]{sep}{desc_ref}))"
                )
                client_specific_formula = (
                    f"=IFERROR(INDEX(ClientRules[[#Data],[Category]]{sep}"
                    f"MATCH(AGGREGATE(15{sep}6{sep}(1*ClientRules[[#Data],[Priority]])/"
                    f"INDEX({cond_expr}{sep}0){sep}1){sep}"
                    f"(1*ClientRules[[#Data],[Priority]]){sep}0))"
                    f"{sep}\"\")"
                )
                ws.cell(row=r, column=client_specific_col).value = client_specific_formula
        if final_col and manual_col and client_specific_col and global_cat_col:
            manual_letter = get_column_letter(manual_col)
            client_letter = get_column_letter(client_specific_col)
            global_letter = get_column_letter(global_cat_col)
            for r in range(2, max_r + 1):
                manual_ref = f"{manual_letter}{r}"
                client_ref = f"{client_letter}{r}"
                global_ref = f"{global_letter}{r}"
                final_category_formula = f'=IF({manual_ref}<>"",{manual_ref},IF({client_ref}<>"",{client_ref},IF({global_ref}<>"",{global_ref},"")))'
                if sep != ",":
                    final_category_formula = final_category_formula.replace(",", sep)
                ws.cell(row=r, column=final_col).value = final_category_formula

        if tn_col:
            ws.column_dimensions[get_column_letter(tn_col)].hidden = True
        if bal_col:
            ws.column_dimensions[get_column_letter(bal_col)].hidden = True

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)

            if ws.column_dimensions[col_letter].hidden:
                continue

            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    continue

                if hasattr(val, "strftime"):
                    try:
                        s = val.strftime("%d/%m/%Y")
                    except Exception:
                        s = str(val)
                elif isinstance(val, (int, float)) and (col_idx == amt_col or col_idx == bal_col):
                    try:
                        s = f"£{float(val):,.2f}"
                    except Exception:
                        s = str(val)
                else:
                    if isinstance(val, str) and val.startswith("="):
                        continue
                    s = str(val)

                if len(s) > max_len:
                    max_len = len(s)

            width = max_len + 2
            if width < 10:
                width = 10
            if width > 60:
                width = 60

            ws.column_dimensions[col_letter].width = width

        if global_cat_col and hasattr(ws, "_cells"):
            for r in range(2, max_r + 1):
                if (r, global_cat_col) in ws._cells:
                    cell = ws._cells[(r, global_cat_col)]
                    if cell.value is None or cell.value == "":
                        del ws._cells[(r, global_cat_col)]

    if enable_categorisation:
        try:
            ensure_folder(LOGS_DIR)
            audit = _audit_xlsx_categorisation(output_path)
            audit_json_path = _write_log_json("categorisation_audit", audit)
            try:
                _write_categorisation_evidence_instructions(audit, audit_json_path)
            except Exception:
                pass
        except Exception:
            details = []
            details.append(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            details.append(f"output_path: {output_path}")
            details.append("Error during post-save categorisation audit:")
            details.append(traceback.format_exc())
            _write_log_text("categorisation_audit_error", "\n".join(details))



# ----------------------------
# Reconciliation
# ----------------------------

def reconcile_statement(parser, pdf_path: str, transactions: list[dict]) -> dict:
    result = {
        "pdf": os.path.basename(pdf_path),
        "start_balance": None,
        "end_balance": None,
        "sum_amounts": None,
        "expected_end": None,
        "difference": None,
        "status": "Not checked",
        # Period metadata for continuity chain + overlap logic
        "period_start": None,
        "period_end": None,
        # Keep txns available for overlap de-dupe plan (no filtering here)
        "transactions": transactions or [],
    }
    # Extract period dates (best-effort). Prefer parser-provided extractor when available.
    # If not found => (None, None)
    try:
        ps = pe = None

        # 1) Prefer parser.extract_statement_period(pdf_path) when present.
        if hasattr(parser, "extract_statement_period"):
            try:
                got = parser.extract_statement_period(pdf_path)
                if isinstance(got, (tuple, list)) and len(got) >= 2:
                    ps, pe = got[0], got[1]
            except Exception:
                ps = pe = None

        # 2) Fallback to Core PDF text extraction (legacy Starling summary line) if parser did not provide.
        if ps is None or pe is None:
            bank_hint = ""
            try:
                bank_hint = getattr(parser, "__name__", "") or ""
            except Exception:
                bank_hint = ""

            ps2, pe2 = extract_statement_period_from_pdf(pdf_path, bank_hint=bank_hint)
            if ps is None:
                ps = ps2
            if pe is None:
                pe = pe2

        result["period_start"] = ps
        result["period_end"] = pe

    except Exception:
        result["period_start"] = None
        result["period_end"] = None

    if not hasattr(parser, "extract_statement_balances"):
        result["status"] = "Not supported by parser"
        return result

    try:
        balances = parser.extract_statement_balances(pdf_path) or {}
        start = balances.get("start_balance")
        end = balances.get("end_balance")
        result["start_balance"] = start
        result["end_balance"] = end

        if start is None or end is None:
            result["status"] = "Statement balances not found"
            return result

        total = 0.0
        for t in transactions or []:
            amt = t.get("Amount")
            if amt is None or amt == "":
                continue
            try:
                total += float(amt)
            except Exception:
                continue

        total = round(total, 2)
        expected_end = round(float(start) + total, 2)
        diff = round(expected_end - float(end), 2)

        result["sum_amounts"] = total
        result["expected_end"] = expected_end
        result["difference"] = diff
        result["status"] = "OK" if abs(diff) <= 0.01 else "Mismatch"
        return result

    except Exception:
        result["status"] = "Error during reconciliation"
        return result


def run_audit_checks_basic(pdf_name: str, transactions: list[dict], start_balance, end_balance) -> dict:
    def _parse_money(v):
        if v is None or v == "":
            return None
        try:
            s = str(v).strip()
        except Exception:
            return None
        if not s:
            return None

        s = s.replace("\u00a0", " ").replace("£", "").replace(",", "")
        s = s.replace("\u2212", "-").replace("\u2013", "-").replace("\u2014", "-").strip()
        s = re.sub(r"\s+", " ", s)

        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1].strip()

        s = re.sub(r"\b(CR|DR|CREDIT|DEBIT)\b\.?$", "", s, flags=re.IGNORECASE).strip()
        s = s.replace(" ", "")
        if not s:
            return None

        try:
            n = float(s)
        except Exception:
            return None
        if neg:
            n = -abs(n)
        return round(n, 2)

    def _is_blank(v):
        if v is None:
            return True
        try:
            return str(v).strip() == ""
        except Exception:
            return True

    row_count = len(transactions or [])
    start_val = _parse_money(start_balance)
    end_val = _parse_money(end_balance)

    parseable_balance_rows = 0
    checked_rows = 0
    missing_or_bad_walk_rows = 0
    mismatch_examples = []
    running = start_val
    last_checked_balance = None

    for idx, txn in enumerate(transactions or [], start=1):
        amt = _parse_money((txn or {}).get("Amount"))
        bal = _parse_money((txn or {}).get("Balance"))

        if bal is not None:
            parseable_balance_rows += 1

        if amt is None or bal is None:
            missing_or_bad_walk_rows += 1
            running = None
            continue

        if running is None:
            running = bal
            last_checked_balance = bal
            continue

        expected = round(running + amt, 2)
        if abs(expected - bal) > 0.01:
            if len(mismatch_examples) < 5:
                mismatch_examples.append(
                    {
                        "row": idx,
                        "expected": expected,
                        "actual": bal,
                    }
                )
        running = bal
        last_checked_balance = bal
        checked_rows += 1

    balance_walk_status = "OK"
    balance_walk_summary = ""
    balance_walk_details = {
        "row_count": row_count,
        "checked_rows": checked_rows,
        "parseable_balance_rows": parseable_balance_rows,
        "missing_or_bad_rows": missing_or_bad_walk_rows,
        "mismatch_examples": mismatch_examples,
        "end_balance_check": None,
    }

    if checked_rows == 0:
        balance_walk_status = "NOT CHECKED"
        reasons = []
        if start_val is None:
            reasons.append("start balance missing/unparseable")
        if parseable_balance_rows == 0:
            reasons.append("no parseable row balances")
        if parseable_balance_rows > 0 and start_val is not None:
            reasons.append("no consecutive parseable amount/balance rows")
        balance_walk_summary = "; ".join(reasons)
    else:
        if mismatch_examples:
            balance_walk_status = "MISMATCH"
            balance_walk_summary = f"{len(mismatch_examples)} row mismatches"
        else:
            balance_walk_summary = f"{checked_rows} rows checked"

        if end_val is not None and last_checked_balance is not None:
            end_diff = round(last_checked_balance - end_val, 2)
            end_ok = abs(end_diff) <= 0.01
            balance_walk_details["end_balance_check"] = {
                "last_checked_balance": last_checked_balance,
                "end_balance": end_val,
                "diff": end_diff,
                "ok": end_ok,
            }
            if not end_ok:
                balance_walk_status = "MISMATCH"
                if balance_walk_summary:
                    balance_walk_summary = balance_walk_summary + "; "
                balance_walk_summary = balance_walk_summary + f"end balance mismatch ({end_diff:+.2f})"

    missing_date = 0
    missing_type = 0
    missing_description = 0
    bad_amount = 0
    bad_balance = 0

    missing_date_rows = []
    missing_type_rows = []
    missing_description_rows = []
    bad_amount_rows = []
    bad_balance_rows = []

    for idx, txn in enumerate(transactions or [], start=1):
        t = txn or {}
        if _is_blank(t.get("Date")):
            missing_date += 1
            if len(missing_date_rows) < 8:
                missing_date_rows.append(idx)
        if _is_blank(t.get("Transaction Type")):
            missing_type += 1
            if len(missing_type_rows) < 8:
                missing_type_rows.append(idx)
        if _is_blank(t.get("Description")):
            missing_description += 1
            if len(missing_description_rows) < 8:
                missing_description_rows.append(idx)
        if _parse_money(t.get("Amount")) is None:
            bad_amount += 1
            if len(bad_amount_rows) < 8:
                bad_amount_rows.append(idx)
        bal_raw = t.get("Balance")
        if (not _is_blank(bal_raw)) and _parse_money(bal_raw) is None:
            bad_balance += 1
            if len(bad_balance_rows) < 8:
                bad_balance_rows.append(idx)

    total_shape_issues = missing_date + missing_type + missing_description + bad_amount + bad_balance
    if total_shape_issues == 0:
        row_shape_status = "OK"
        row_shape_summary = "all required fields parseable"
    else:
        row_shape_status = "WARN"
        row_shape_summary = (
            f"missing Date {missing_date}, Type {missing_type}, Description {missing_description}, "
            f"bad Amount {bad_amount}, bad Balance {bad_balance}"
        )

    row_shape_details = {
        "row_count": row_count,
        "missing_date": {"count": missing_date, "rows": missing_date_rows},
        "missing_type": {"count": missing_type, "rows": missing_type_rows},
        "missing_description": {"count": missing_description, "rows": missing_description_rows},
        "bad_amount": {"count": bad_amount, "rows": bad_amount_rows},
        "bad_balance": {"count": bad_balance, "rows": bad_balance_rows},
    }

    if balance_walk_status == "MISMATCH":
        overall_status = "MISMATCH"
    elif row_shape_status == "WARN":
        overall_status = "WARN"
    else:
        overall_status = "OK"

    return {
        "pdf": pdf_name,
        "balance_walk_status": balance_walk_status,
        "balance_walk_summary": balance_walk_summary,
        "balance_walk_details": balance_walk_details,
        "row_shape_status": row_shape_status,
        "row_shape_summary": row_shape_summary,
        "row_shape_details": row_shape_details,
        "status": overall_status,
    }


def _overlap_dedupe_continuity_resolution(
    a_txns: list[dict] | None,
    b_txns: list[dict] | None,
    a_period_start,
    a_period_end,
    b_period_start,
    b_period_end,
    a_end_balance,
    b_start_balance,
    logger=None,
) -> dict:
    """Overlap-aware continuity resolution: produce a de-dupe plan (do not filter here).

    Enforces:
    - overlap exists if B.period_start <= A.period_end
    - overlap window = [B.period_start, min(A.period_end, B.period_end)]
    - duplicates detected only within overlap window
    - duplicates removed from B only
    - signature rules:
        * if BOTH matched rows have balance -> require balance signature
        * fallback (no balance) only if either side is missing balance
    - safety gate:
        diff = A.end_balance - B.start_balance
        apply only if abs(diff - dupe_sum) <= 0.01
    """

    log_lines: list[str] = []
    log_fields: list[dict] = []

    def _log(msg: str, **fields):
        log_lines.append(msg)
        if fields:
            log_fields.append(fields)
        try:
            if callable(logger):
                logger(msg, fields if fields else None)
        except Exception:
            pass

    def _coerce_date(v):
        if v is None or v == "":
            return None
        try:
            if hasattr(v, "to_pydatetime"):
                v = v.to_pydatetime()
        except Exception:
            pass
        try:
            if isinstance(v, datetime):
                return v.date()
        except Exception:
            pass
        if isinstance(v, date):
            return v
        try:
            if hasattr(v, "date"):
                return v.date()
        except Exception:
            pass
        return None

    def _norm_text(v):
        if v is None:
            return ""
        s = str(v).upper()
        s = " ".join(s.split())
        return s.strip()

    def _money_2dp(v):
        if v is None or v == "":
            return None
        try:
            return round(float(v), 2)
        except Exception:
            return None

    def _fallback_sig(tx: dict):
        return (
            _coerce_date(tx.get("Date")),
            _money_2dp(tx.get("Amount")),
            _norm_text(tx.get("Transaction Type")),
            _norm_text(tx.get("Description")),
        )

    def _bal_sig(tx: dict):
        return (
            _coerce_date(tx.get("Date")),
            _money_2dp(tx.get("Amount")),
            _money_2dp(tx.get("Balance")),
            _norm_text(tx.get("Transaction Type")),
            _norm_text(tx.get("Description")),
        )

    a_txns = a_txns or []
    b_txns = b_txns or []

    a_ps = _coerce_date(a_period_start)
    a_pe = _coerce_date(a_period_end)
    b_ps = _coerce_date(b_period_start)
    b_pe = _coerce_date(b_period_end)

    if not (a_ps and a_pe and b_ps and b_pe):
        return {
            "applied": False,
            "removed_count": 0,
            "dupe_sum": 0.0,
            "b_remove_indices": [],
            "status": "Mismatch",
            "log_lines": log_lines,
            "log_fields": log_fields,
        }

    if b_ps > a_pe:
        return {
            "applied": False,
            "removed_count": 0,
            "dupe_sum": 0.0,
            "b_remove_indices": [],
            "status": "Mismatch",
            "log_lines": log_lines,
            "log_fields": log_fields,
        }

    overlap_start = b_ps
    overlap_end = min(a_pe, b_pe)

    a_bal = Counter()
    a_fb_all = Counter()
    a_fb_no_bal = Counter()

    for tx in a_txns:
        if not isinstance(tx, dict):
            continue
        d = _coerce_date(tx.get("Date"))
        if not d or d < overlap_start or d > overlap_end:
            continue

        fs = _fallback_sig(tx)
        if fs[0] is None or fs[1] is None:
            continue

        a_fb_all[fs] += 1

        if _money_2dp(tx.get("Balance")) is None:
            a_fb_no_bal[fs] += 1
        else:
            bs = _bal_sig(tx)
            if bs[0] is not None and bs[1] is not None and bs[2] is not None:
                a_bal[bs] += 1

    if not a_fb_all and not a_bal:
        return {
            "applied": False,
            "removed_count": 0,
            "dupe_sum": 0.0,
            "b_remove_indices": [],
            "status": "Mismatch",
            "log_lines": log_lines,
            "log_fields": log_fields,
        }

    b_remove_indices: list[int] = []
    dupe_sum = 0.0

    for idx, tx in enumerate(b_txns):
        if not isinstance(tx, dict):
            continue
        d = _coerce_date(tx.get("Date"))
        if not d or d < overlap_start or d > overlap_end:
            continue

        amt = _money_2dp(tx.get("Amount"))
        if amt is None:
            continue

        b_has_bal = _money_2dp(tx.get("Balance")) is not None
        matched = False

        if b_has_bal:
            bs = _bal_sig(tx)
            if bs[0] is not None and bs[1] is not None and bs[2] is not None and a_bal.get(bs, 0) > 0:
                a_bal[bs] -= 1
                if a_bal[bs] <= 0:
                    del a_bal[bs]

                fs = _fallback_sig(tx)
                if a_fb_all.get(fs, 0) > 0:
                    a_fb_all[fs] -= 1
                    if a_fb_all[fs] <= 0:
                        del a_fb_all[fs]

                matched = True

            if not matched:
                fs = _fallback_sig(tx)
                if fs[0] is not None and fs[1] is not None and a_fb_no_bal.get(fs, 0) > 0:
                    a_fb_no_bal[fs] -= 1
                    if a_fb_no_bal[fs] <= 0:
                        del a_fb_no_bal[fs]

                    if a_fb_all.get(fs, 0) > 0:
                        a_fb_all[fs] -= 1
                        if a_fb_all[fs] <= 0:
                            del a_fb_all[fs]

                    matched = True

        else:
            fs = _fallback_sig(tx)
            if fs[0] is not None and fs[1] is not None and a_fb_all.get(fs, 0) > 0:
                a_fb_all[fs] -= 1
                if a_fb_all[fs] <= 0:
                    del a_fb_all[fs]

                if a_fb_no_bal.get(fs, 0) > 0:
                    a_fb_no_bal[fs] -= 1
                    if a_fb_no_bal[fs] <= 0:
                        del a_fb_no_bal[fs]

                matched = True

        if matched:
            b_remove_indices.append(idx)
            dupe_sum = round(dupe_sum + amt, 2)

    removed_count = len(b_remove_indices)

    a_end = _money_2dp(a_end_balance)
    b_start = _money_2dp(b_start_balance)

    if a_end is None or b_start is None:
        return {
            "applied": False,
            "removed_count": 0,
            "dupe_sum": dupe_sum,
            "b_remove_indices": [],
            "status": "Mismatch",
            "log_lines": log_lines,
            "log_fields": log_fields,
        }

    diff = round(a_end - b_start, 2)

    _log(
        "Overlap resolution attempt",
        prev_end=a_end,
        next_start=b_start,
        diff=diff,
        overlap_start=str(overlap_start),
        overlap_end=str(overlap_end),
        removed_count=removed_count,
        dupe_sum=dupe_sum,
    )

    if removed_count <= 0:
        return {
            "applied": False,
            "removed_count": 0,
            "dupe_sum": 0.0,
            "b_remove_indices": [],
            "status": "Mismatch",
            "log_lines": log_lines,
            "log_fields": log_fields,
        }

    if abs(diff - dupe_sum) <= 0.01:
        return {
            "applied": True,
            "removed_count": removed_count,
            "dupe_sum": dupe_sum,
            "b_remove_indices": b_remove_indices,
            "status": f"OK (overlap resolved, removed {removed_count} duplicates, sum {_fmt_money(dupe_sum)})",
            "log_lines": log_lines,
            "log_fields": log_fields,
        }

    _log("Overlap safety gate failed; not applying", diff=diff, dupe_sum=dupe_sum, removed_count=removed_count)

    return {
        "applied": False,
        "removed_count": 0,
        "dupe_sum": dupe_sum,
        "b_remove_indices": [],
        "status": "Mismatch",
        "log_lines": log_lines,
        "log_fields": log_fields,
    }



def compute_statement_continuity(recon_results: list[dict]) -> list[dict]:
    if not recon_results or len(recon_results) < 2:
        return []

    import pandas as pd

    def _safe_date(d):
        if d is None:
            return None
        try:
            if hasattr(d, "to_pydatetime"):
                d = d.to_pydatetime()
        except Exception:
            pass
        try:
            if isinstance(d, datetime):
                return d.date()
        except Exception:
            pass
        return d

    def _sort_key_idx(i: int):
        r = recon_results[i]
        d = _safe_date(r.get("date_min"))
        if d is None:
            return (1, datetime.max.date(), str(r.get("pdf", "")))
        return (0, d, str(r.get("pdf", "")))

    def _to_money_key(v):
        """Normalize statement-level balances for continuity.

        Treat 0.00 as a valid value. Only None/blank/NaN are considered missing.
        Robustly parse common UK statement formats, e.g.:
          "£0.00", "1,234.56", "1,234.56 CR", "(12.34)", unicode minus.
        """
        if v is None or v == "":
            return None

        # Handle pandas/NumPy NaN
        try:
            if isinstance(v, float) and pd.isna(v):
                return None
        except Exception:
            pass

        try:
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    return None

                # Normalize whitespace + currency/grouping
                s = s.replace("\u00A0", " ")  # NBSP
                s = s.replace("£", "")
                s = s.replace(",", "")

                # Normalize unicode minus/dash
                s = s.replace("−", "-").replace("–", "-").replace("—", "-")

                # Parentheses -> negative
                if s.startswith("(") and s.endswith(")"):
                    s = "-" + s[1:-1].strip()

                # Remove trailing markers (CR/DR etc.)
                s = re.sub(r"\s*(CR|DR|CREDIT|DEBIT)\s*$", "", s, flags=re.IGNORECASE)

                # Extract first numeric token (keeps leading sign)
                m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
                if not m:
                    return None
                f = float(m.group(0))
            else:
                f = float(v)

            if pd.isna(f):
                return None
            return round(f, 2)
        except Exception:
            return None

    def _start_for_chain(r):
        """Start balance used for continuity linking.

        Per requirement: use statement-level start_balance (from reconciliation).
        Only fall back to continuity_start_balance if statement start_balance is missing.
        """
        v = r.get("start_balance")
        if v is None or v == "":
            v = r.get("continuity_start_balance")
        return v

    def _period_start_idx(i: int):
        return _safe_date(recon_results[i].get("period_start"))

    def _period_end_idx(i: int):
        return _safe_date(recon_results[i].get("period_end"))

    n = len(recon_results)

    start_map: dict[float, list[int]] = {}
    end_map: dict[float, list[int]] = {}

    for i, r in enumerate(recon_results):
        sk = _to_money_key(_start_for_chain(r))
        ek = _to_money_key(r.get("end_balance"))
        if sk is not None:
            start_map.setdefault(sk, []).append(i)
        if ek is not None:
            end_map.setdefault(ek, []).append(i)

    # Head selection:
    # If we have any reliable period_start values, always start from the earliest period_start.
    # This prevents the chain from starting mid-way and later "jumping backwards" when appending remaining statements.
    period_heads = [i for i in range(n) if isinstance(_period_start_idx(i), date)]
    if period_heads:
        head_idx = sorted(period_heads, key=lambda i: (_period_start_idx(i), _sort_key_idx(i)))[0]
    else:
        head_candidates: list[int] = []
        for i, r in enumerate(recon_results):
            sk = _to_money_key(_start_for_chain(r))
            ek = _to_money_key(r.get("end_balance"))
            if sk is None or ek is None:
                continue
            if sk not in end_map:
                head_candidates.append(i)

        if head_candidates:
            head_idx = sorted(head_candidates, key=_sort_key_idx)[0]
        else:
            head_idx = sorted(range(n), key=_sort_key_idx)[0]

    visited: set[int] = set()
    ordered_idx: list[int] = []

    chain_link_logs: dict[tuple[int, int], dict] = {}

    def _log_link(a_idx: int, b_idx: int, note: str, **extra):
        chain_link_logs[(a_idx, b_idx)] = {
            "a_pdf": recon_results[a_idx].get("pdf"),
            "b_pdf": recon_results[b_idx].get("pdf") if b_idx is not None else None,
            "a_period_start": str(_period_start_idx(a_idx)) if _period_start_idx(a_idx) else None,
            "a_period_end": str(_period_end_idx(a_idx)) if _period_end_idx(a_idx) else None,
            "b_period_start": str(_period_start_idx(b_idx)) if (b_idx is not None and _period_start_idx(b_idx)) else None,
            "b_period_end": str(_period_end_idx(b_idx)) if (b_idx is not None and _period_end_idx(b_idx)) else None,
            "note": note,
        }
        if extra:
            chain_link_logs[(a_idx, b_idx)].update(extra)

    def _choose_next_chronological(cur_idx: int) -> int | None:
        """Fallback when balance-linking fails or is unusable.

        When A has period_start, pick the next unvisited statement with the smallest period_start
        that is >= A.period_start (never jump backwards). Prefer the closest to A.period_end.
        """
        a_ps = _period_start_idx(cur_idx)
        a_pe = _period_end_idx(cur_idx)

        # If A period_start isn't known, we cannot enforce a chronology gate.
        if not isinstance(a_ps, date):
            return None

        candidates = []
        for j in range(n):
            if j in visited:
                continue
            b_ps = _period_start_idx(j)
            if not isinstance(b_ps, date):
                continue
            if b_ps < a_ps:
                continue
            gap = None
            if isinstance(a_pe, date):
                try:
                    gap = abs((b_ps - a_pe).days)
                except Exception:
                    gap = None
            candidates.append((gap if gap is not None else 10**9, b_ps, _sort_key_idx(j), j))

        if not candidates:
            return None

        candidates.sort(key=lambda t: (t[0], t[1], t[2]))
        return candidates[0][3]

    cur = head_idx
    while cur not in visited:
        visited.add(cur)
        ordered_idx.append(cur)

        cur_end = _to_money_key(recon_results[cur].get("end_balance"))
        if cur_end is None:
            break

        # Balance-link candidates
        nxt_candidates = [j for j in start_map.get(cur_end, []) if j not in visited]

        a_ps = _period_start_idx(cur)
        a_pe = _period_end_idx(cur)

        # Enforce chronology gate on ALL selection paths when periods are known.
        chrono_can_apply = isinstance(a_ps, date)

        if nxt_candidates:
            known_b: list[int] = []
            unknown_b: list[int] = []
            pass_b: list[int] = []
            fail_b: list[int] = []

            for j in nxt_candidates:
                b_ps = _period_start_idx(j)
                if not (isinstance(a_ps, date) and isinstance(b_ps, date)):
                    unknown_b.append(j)
                    continue
                known_b.append(j)
                if b_ps >= a_ps:
                    pass_b.append(j)
                else:
                    fail_b.append(j)

            # Apply gate: if both are known and B is earlier than A, disallow.
            allowed = [j for j in nxt_candidates if j not in fail_b]

            chrono_applied = bool(known_b) and chrono_can_apply
            if not chrono_can_apply or not known_b:
                chrono_note = "chronology gate skipped (missing period dates)"
            else:
                chrono_note = "Chronology gate applied."

            if pass_b:
                # Prefer earliest B.period_start (>= A.period_start)
                chosen = sorted(pass_b, key=lambda j: (_period_start_idx(j), _sort_key_idx(j)))[0]
                _log_link(
                    cur,
                    chosen,
                    "Balance link used; " + "Chronology gate applied; chose earliest B.period_start." if chrono_applied else "Balance link used; chronology gate skipped (missing period dates).",
                    candidates_total=len(nxt_candidates),
                    candidates_known=len(known_b),
                    candidates_pass=len(pass_b),
                    candidates_fail=len(fail_b),
                    candidates_unknown=len(unknown_b),
                    chronology_gate_applied=chrono_applied,
                )
                cur = chosen
                continue

            # No passing candidates. If we can apply chronology, do NOT pick a backward/invalid candidate.
            if chrono_can_apply:
                fb = _choose_next_chronological(cur)
                if fb is not None:
                    _log_link(
                        cur,
                        fb,
                        "Balance link ambiguous/no forward candidate; fallback used: chronological next (never backwards).",
                        candidates_total=len(nxt_candidates),
                        candidates_known=len(known_b),
                        candidates_pass=len(pass_b),
                        candidates_fail=len(fail_b),
                        candidates_unknown=len(unknown_b),
                        chronology_gate_applied=True,
                    )
                    cur = fb
                    continue
                else:
                    # No forward chronological candidate exists -> terminate chain.
                    _log_link(
                        cur,
                        cur,
                        "Balance link ambiguous and no chronological forward candidate exists; chain terminated.",
                        candidates_total=len(nxt_candidates),
                        candidates_known=len(known_b),
                        candidates_pass=len(pass_b),
                        candidates_fail=len(fail_b),
                        candidates_unknown=len(unknown_b),
                        chronology_gate_applied=True,
                    )
                    break

            # Chronology cannot apply; keep previous behaviour but log it.
            chosen = sorted(allowed or nxt_candidates, key=_sort_key_idx)[0]
            _log_link(
                cur,
                chosen,
                "Balance link used; chronology gate skipped (missing period dates).",
                candidates_total=len(nxt_candidates),
                candidates_known=len(known_b),
                candidates_pass=len(pass_b),
                candidates_fail=len(fail_b),
                candidates_unknown=len(unknown_b),
                chronology_gate_applied=False,
            )
            cur = chosen
            continue

        # No balance-link candidate exists.
        if chrono_can_apply:
            fb = _choose_next_chronological(cur)
            if fb is not None:
                _log_link(
                    cur,
                    fb,
                    "Balance link failed (no candidate); fallback used: chronological next (never backwards).",
                    chronology_gate_applied=True,
                )
                cur = fb
                continue
            else:
                _log_link(
                    cur,
                    cur,
                    "Balance link failed and no chronological forward candidate exists; chain terminated.",
                    chronology_gate_applied=True,
                )
                break

        # If we cannot apply chronology, terminate the chain (prevents arbitrary mid-way ordering).
        _log_link(
            cur,
            cur,
            "Balance link failed; chronology gate skipped (missing period dates); chain terminated.",
            chronology_gate_applied=False,
        )
        break

    # Append any remaining statements.
    # If periods exist, append them in chronological order by period_start.
    remaining = [i for i in range(n) if i not in visited]
    if remaining:
        rem_with_ps = [i for i in remaining if isinstance(_period_start_idx(i), date)]
        rem_without_ps = [i for i in remaining if i not in rem_with_ps]

        ordered_idx.extend(sorted(rem_with_ps, key=lambda i: (_period_start_idx(i), _sort_key_idx(i))))
        ordered_idx.extend(sorted(rem_without_ps, key=_sort_key_idx))

    links: list[dict] = []
    for pos in range(len(ordered_idx) - 1):
        a_idx = ordered_idx[pos]
        b_idx = ordered_idx[pos + 1]
        a = recon_results[a_idx]
        b = recon_results[b_idx]

        prev_end_raw = a.get("end_balance")
        next_start_raw = _start_for_chain(b)

        prev_end = _to_money_key(prev_end_raw)
        next_start = _to_money_key(next_start_raw)

        prev_date_max = _safe_date(a.get("date_max"))
        next_date_min = _safe_date(b.get("date_min"))

        missing_from = None
        missing_to = None
        try:
            if isinstance(prev_date_max, date) and isinstance(next_date_min, date):
                _mf = prev_date_max + timedelta(days=1)
                _mt = next_date_min - timedelta(days=1)
                if _mf <= _mt:
                    missing_from = _mf
                    missing_to = _mt
        except Exception:
            missing_from = None
            missing_to = None

        chain_log = chain_link_logs.get((a_idx, b_idx)) or {}

        link = {
            "prev_pdf": a.get("pdf"),
            "next_pdf": b.get("pdf"),
            # Debug: raw statement-level balances used for continuity linking
            "prev_end_raw": prev_end_raw,
            "next_start_raw": next_start_raw,
            "chain_candidates_total": chain_log.get("candidates_total"),
            "chain_candidates_known_period_start": chain_log.get("candidates_known"),
            "chain_candidates_chrono_pass": chain_log.get("candidates_pass"),
            "chain_candidates_chrono_fail": chain_log.get("candidates_fail"),
            "chain_candidates_chrono_unknown": chain_log.get("candidates_unknown"),
            "chronology_gate_applied": chain_log.get("chronology_gate_applied"),
            "chronology_gate_note": chain_log.get("note") or chain_log.get("note"),
            "prev_end": prev_end,
            "next_start": next_start,
            "prev_date_max": prev_date_max,
            "next_date_min": next_date_min,
            "missing_from": missing_from,
            "missing_to": missing_to,
            "diff": None,
            "status": "Not checked",
        }

        if prev_end is None or next_start is None:
            links.append(link)
            continue

        try:
            diff = round(float(prev_end) - float(next_start), 2)
            link["diff"] = diff
            link["status"] = "OK" if abs(diff) <= 0.01 else "Mismatch"
        except Exception:
            link["status"] = "Not checked"

        # Overlap-aware continuity resolution (produce a de-dupe plan)
        link["applied_overlap_resolution"] = False
        link["overlap_window"] = None
        link["duplicates_to_remove_from_B"] = []
        link["removed_count"] = 0
        link["dupe_sum"] = 0.0
        link["display_status"] = link.get("status")
        link["overlap_log_lines"] = []
        link["overlap_log_fields"] = []

        # Only attempt overlap resolution when first-pass mismatches AND both A/B have period_start/period_end.
        if link.get("status") == "Mismatch":
            a_txns = a.get("transactions") or a.get("txns")
            b_txns = b.get("transactions") or b.get("txns")

            a_ps = a.get("period_start")
            a_pe = a.get("period_end")
            b_ps = b.get("period_start")
            b_pe = b.get("period_end")

            have_periods = (
                a_ps is not None
                and a_pe is not None
                and b_ps is not None
                and b_pe is not None
            )

            if not have_periods:
                link["overlap_log_lines"].append(
                    "Overlap resolution not attempted (missing period_start/period_end on A or B)."
                )
                link["overlap_log_fields"].append(
                    {
                        "prev_pdf": a.get("pdf"),
                        "next_pdf": b.get("pdf"),
                        "note": "missing period_start/period_end",
                    }
                )
            else:
                _a_pe = _safe_date(a_pe)
                _b_ps = _safe_date(b_ps)
                _b_pe = _safe_date(b_pe)
                # Boundary-day overlap is valid (<=).
                if _a_pe and _b_ps and _b_pe and (_b_ps <= _a_pe):
                    overlap_start = _b_ps
                    overlap_end = min(_a_pe, _b_pe)
                    link["overlap_window"] = {"start": overlap_start, "end": overlap_end}

                if isinstance(a_txns, list) and isinstance(b_txns, list):
                    res = _overlap_dedupe_continuity_resolution(
                        a_txns=a_txns,
                        b_txns=b_txns,
                        a_period_start=a_ps,
                        a_period_end=a_pe,
                        b_period_start=b_ps,
                        b_period_end=b_pe,
                        a_end_balance=prev_end,
                        b_start_balance=next_start,
                        logger=None,
                    )

                    link["applied_overlap_resolution"] = bool(res.get("applied"))
                    link["duplicates_to_remove_from_B"] = res.get("b_remove_indices") or []
                    link["removed_count"] = int(res.get("removed_count") or 0)
                    try:
                        link["dupe_sum"] = float(res.get("dupe_sum") or 0.0)
                    except Exception:
                        link["dupe_sum"] = res.get("dupe_sum")

                    link["overlap_log_lines"] = res.get("log_lines") or []
                    link["overlap_log_fields"] = res.get("log_fields") or []

                    if res.get("applied"):
                        link["display_status"] = res.get("status") or link.get("status")
                        link["status"] = link["display_status"]
                    else:
                        link["display_status"] = link.get("status")
                else:
                    link["overlap_log_lines"].append(
                        "Overlap resolution not attempted (missing transactions list on A or B)."
                    )
                    link["overlap_log_fields"].append(
                        {
                            "prev_pdf": a.get("pdf"),
                            "next_pdf": b.get("pdf"),
                            "note": "missing transactions list",
                        }
                    )

        links.append(link)

    return links


def _fmt_money(v):
    try:
        if v is None or v == "":
            return ""
        return f"£{float(v):,.2f}"
    except Exception:
        return str(v)


def show_reconciliation_popup(
    parent,
    output_path: str,
    recon_results: list[dict],
    coverage_period: str = "",
    continuity_results: list[dict] | None = None,
    pre_save: bool = False,
    open_log_folder_callback=None,
):
    continuity_results = continuity_results or []

    any_recon_warn = any((r.get("status") or "") != "OK" for r in (recon_results or []))
    any_cont_warn = any((r.get("status") or "") != "OK" for r in (continuity_results or []))
    any_warn = any_recon_warn or any_cont_warn

    def _norm_date(v):
        try:
            if v is None or v == "":
                return None
            if hasattr(v, "to_pydatetime"):
                v = v.to_pydatetime()
            if hasattr(v, "date") and not isinstance(v, date):
                return v.date()
            return v
        except Exception:
            return None

    def _recon_sort_key(r):
        d = _norm_date(r.get("date_min"))
        d = d or _norm_date(r.get("date_max"))
        return (d or date.min, str(r.get("pdf") or ""))

    recon_results = sorted(list(recon_results or []), key=_recon_sort_key)

    win = tk.Toplevel(parent)
    win.transient(parent)
    win.grab_set()

    win.title("Reconciliation warning" if any_warn else "Success")
    win.geometry("820x560")

    outer = ttk.Frame(win, padding=14)
    outer.pack(fill="both", expand=True)

    icon = "✖" if any_warn else "✔"
    icon_color = "#b00020" if any_warn else "#0b6e0b"

    head = ttk.Frame(outer)
    head.pack(fill="x")

    ttk.Label(head, text=icon, foreground=icon_color, font=("Segoe UI", 18, "bold")).pack(side="left")

    title_text = "Checks completed with warnings" if any_warn else (
        "Checks completed" if pre_save else "Excel created successfully"
    )
    ttk.Label(head, text=title_text, font=("Segoe UI", 13, "bold")).pack(side="left", padx=(10, 0))

    display_path = output_path

    path_row = ttk.Frame(outer)
    path_row.pack(fill="x", pady=(10, 0))

    ttk.Label(path_row, text="Output:").pack(side="left")
    ttk.Label(path_row, text=display_path, foreground="#333").pack(side="left", padx=(6, 0))

    txt = tk.Text(outer, height=22, wrap="word")
    txt.pack(fill="both", expand=True, pady=(12, 0))

    txt.tag_configure("section", font=("Segoe UI", 10, "bold"))
    txt.tag_configure("ok", foreground="#0b6e0b")
    txt.tag_configure("bad", foreground="#b00020")
    txt.tag_configure("warn", foreground="#8a6d3b")
    txt.tag_configure("info", foreground="#333")

    lines_for_clipboard: list[str] = []

    if coverage_period:
        if any_warn:
            line = (
                f"The bank statements cover the period from {coverage_period} "
                "(however, some checks could not be completed or warnings were detected — see below).\n\n"
            )
        else:
            line = f"The bank statements cover the period from {coverage_period}.\n\n"

        txt.insert("end", line, "info")
        lines_for_clipboard.append(line.strip())

    # FIX: ensure string literals are correctly terminated
    txt.insert("end", "Reconciliation check:\n", "section")
    for r in recon_results:
        status = (r.get("status") or "").strip()
        pdf = r.get("pdf") or ""

        ps = r.get("period_start")
        pe = r.get("period_end")
        if ps and pe and hasattr(ps, "strftime") and hasattr(pe, "strftime"):
            period_str = f"Period: {ps.strftime('%d/%m/%Y')} - {pe.strftime('%d/%m/%Y')}"
        else:
            period_str = "Period: None"

        if status == "OK":
            msg = (
                f"OK: {pdf} "
                f"(Start {_fmt_money(r.get('start_balance'))}, Net {_fmt_money(r.get('sum_amounts'))}, End {_fmt_money(r.get('end_balance'))})"
            )
            tag = "ok"
        elif status == "Mismatch":
            msg = (
                f"MISMATCH: {pdf} "
                f"(Start {_fmt_money(r.get('start_balance'))}, Net {_fmt_money(r.get('sum_amounts'))}, End {_fmt_money(r.get('end_balance'))}, Diff {_fmt_money(r.get('difference'))})"
            )
            tag = "bad"
        elif status == "Statement balances not found":
            msg = f"NOT CHECKED: {pdf} (statement balances not found)"
            tag = "warn"
        elif status == "Not supported by parser":
            msg = f"NOT CHECKED: {pdf} (balances not supported by parser)"
            tag = "warn"
        else:
            msg = f"NOT CHECKED: {pdf} ({status or 'not checked'})"
            tag = "warn"

        msg = msg + f" | {period_str}"

        txt.insert("end", msg + "\n", tag)
        lines_for_clipboard.append(msg)

    txt.insert("end", "\n")
    lines_for_clipboard.append("")

    if continuity_results:
        txt.insert("end", "Statement continuity check:\n", "section")

        for c in continuity_results:
            # Prefer display_status (e.g. "OK (overlap resolved, …)") over the base status.
            status = (c.get("display_status") or c.get("status") or "").strip()
            prev_pdf = c.get("prev_pdf") or ""
            next_pdf = c.get("next_pdf") or ""

            prev_end = c.get("prev_end")
            next_start = c.get("next_start")

            # If balances truly missing, mark not checked.
            if prev_end is None or next_start is None:
                msg = (
                    f"NOT CHECKED: {prev_pdf} -> {next_pdf} (balances not found) "
                    f"[prev_end_raw={c.get('prev_end_raw')!r}, next_start_raw={c.get('next_start_raw')!r}, "
                    f"prev_end={c.get('prev_end')!r}, next_start={c.get('next_start')!r}]"
                )
                tag = "warn"

            # Treat any status starting with OK as OK (handles overlap-resolved strings).
            elif status.upper().startswith("OK"):
                msg = (
                    f"OK: {prev_pdf} -> {next_pdf} "
                    f"(End {_fmt_money(prev_end)} matches Start {_fmt_money(next_start)})"
                )
                # If it’s an enriched OK status, include it.
                if status != "OK":
                    msg = msg.replace("OK:", f"{status}:")
                tag = "ok"

            # Treat any status starting with MISMATCH as mismatch.
            elif status.upper().startswith("MISMATCH"):
                missing = ""
                try:
                    mf = c.get("missing_from")
                    mt = c.get("missing_to")
                    if mf and mt and hasattr(mf, "strftime") and hasattr(mt, "strftime"):
                        missing = f" Missing: {mf.strftime('%d/%m/%Y')} - {mt.strftime('%d/%m/%Y')}"
                except Exception:
                    missing = ""

                msg = (
                    f"MISMATCH: {prev_pdf} -> {next_pdf} "
                    f"(End {_fmt_money(prev_end)} vs Start {_fmt_money(next_start)}, Diff {_fmt_money(c.get('diff'))}){missing}"
                )
                tag = "bad"

            else:
                # Unexpected status string; show it rather than incorrectly claiming balances are missing.
                msg = (
                    f"{status or 'NOT CHECKED'}: {prev_pdf} -> {next_pdf} "
                    f"(End {_fmt_money(prev_end)} vs Start {_fmt_money(next_start)}, Diff {_fmt_money(c.get('diff'))})"
                )
                tag = "warn"

            # FIX: close the string properly and include newline escape.
            txt.insert("end", msg + "\n", tag)
            lines_for_clipboard.append(msg)

        txt.insert("end", "\n")
        lines_for_clipboard.append("")

    btns = ttk.Frame(win, padding=(14, 0, 14, 14))
    btns.pack(fill="x")

    result = {"proceed": True}

    def _close(proceed: bool):
        result["proceed"] = proceed
        win.destroy()

    if any_warn and callable(open_log_folder_callback):

        def _open_log_folder():
            try:
                open_log_folder_callback()
            except Exception as e:
                messagebox.showerror("Open log folder error", str(e))

        ttk.Button(btns, text="Open Log Folder", command=_open_log_folder).pack(side="left", padx=(10, 0))

    if pre_save:
        ttk.Button(btns, text="Cancel", command=lambda: _close(False)).pack(side="right")
        ttk.Button(btns, text="OK", command=lambda: _close(True)).pack(side="right", padx=(0, 8))
        win.protocol("WM_DELETE_WINDOW", lambda: _close(True))
    else:
        ttk.Button(btns, text="OK", command=lambda: _close(True)).pack(side="right")
        win.protocol("WM_DELETE_WINDOW", lambda: _close(True))

    parent.wait_window(win)
    return bool(result["proceed"])


# ----------------------------
# Self-tests (do not run in normal GUI usage)
# ----------------------------

def _run_self_tests() -> None:
    # parse_dnd_event_files: braces + spaces
    s = '{C:\\A Folder\\file one.pdf} {C:\\B\\two.pdf}'
    got = parse_dnd_event_files(s)
    assert got == [r'C:\\A Folder\\file one.pdf', r'C:\\B\\two.pdf'], got

    s2 = 'C:\\x\\a.pdf C:\\y\\b.pdf'
    got2 = parse_dnd_event_files(s2)
    assert got2 == [r'C:\\x\\a.pdf', r'C:\\y\\b.pdf'], got2

    rr = [
        {
            "pdf": "2.pdf",
            "start_balance": 100.0,
            "end_balance": 150.0,
            "continuity_start_balance": 100.0,
            "date_min": date(2024, 2, 1),
            "date_max": date(2024, 2, 28),
        },
        {
            "pdf": "1.pdf",
            "start_balance": 50.0,
            "end_balance": 100.0,
            "continuity_start_balance": 50.0,
            "date_min": date(2024, 1, 1),
            "date_max": date(2024, 1, 31),
        },
    ]
    links = compute_statement_continuity(rr)
    assert len(links) == 1, links
    assert links[0]["prev_pdf"] == "1.pdf", links
    assert links[0]["next_pdf"] == "2.pdf", links
    assert links[0]["status"] == "OK", links

    # Chronology head selection: if period_start exists, start at earliest period_start
    rr_ps = [
        {
            "pdf": "6.pdf",
            "start_balance": 0.0,
            "end_balance": 0.0,
            "continuity_start_balance": 0.0,
            "date_min": date(2024, 10, 24),
            "date_max": date(2024, 11, 22),
            "period_start": date(2024, 10, 24),
            "period_end": date(2024, 11, 22),
        },
        {
            "pdf": "1.pdf",
            "start_balance": 0.0,
            "end_balance": 0.0,
            "continuity_start_balance": 0.0,
            "date_min": date(2024, 6, 3),
            "date_max": date(2024, 6, 22),
            "period_start": date(2024, 6, 3),
            "period_end": date(2024, 6, 22),
        },
        {
            "pdf": "3.pdf",
            "start_balance": 0.0,
            "end_balance": 0.0,
            "continuity_start_balance": 0.0,
            "date_min": date(2024, 7, 1),
            "date_max": date(2024, 7, 31),
            "period_start": date(2024, 7, 1),
            "period_end": date(2024, 7, 31),
        },
    ]
    links_ps = compute_statement_continuity(rr_ps)
    assert links_ps[0]["prev_pdf"] == "1.pdf", links_ps
    assert links_ps[0]["next_pdf"] == "3.pdf", links_ps

    # Balance-link failure fallback should never jump backwards when periods exist
    rr_fb = [
        {
            "pdf": "A.pdf",
            "start_balance": 10.0,
            "end_balance": 20.0,
            "continuity_start_balance": 10.0,
            "period_start": date(2024, 1, 1),
            "period_end": date(2024, 1, 31),
            "date_min": date(2024, 1, 1),
            "date_max": date(2024, 1, 31),
        },
        {
            "pdf": "B.pdf",
            "start_balance": 999.0,
            "end_balance": 1111.0,
            "continuity_start_balance": 999.0,
            "period_start": date(2024, 2, 1),
            "period_end": date(2024, 2, 29),
            "date_min": date(2024, 2, 1),
            "date_max": date(2024, 2, 29),
        },
        {
            "pdf": "C.pdf",
            "start_balance": 5.0,
            "end_balance": 10.0,
            "continuity_start_balance": 5.0,
            "period_start": date(2023, 12, 1),
            "period_end": date(2023, 12, 31),
            "date_min": date(2023, 12, 1),
            "date_max": date(2023, 12, 31),
        },
    ]
    links_fb = compute_statement_continuity(rr_fb)
    # Must start at earliest period_start (C), not jump backwards later
    assert links_fb[0]["prev_pdf"] == "C.pdf", links_fb
    assert links_fb[0]["next_pdf"] in ("A.pdf", "B.pdf"), links_fb

    rr_bad = [
        {
            "pdf": "A.pdf",
            "start_balance": 10.0,
            "end_balance": 20.0,
            "continuity_start_balance": 10.0,
            "date_min": date(2024, 1, 1),
            "date_max": date(2024, 1, 31),
        },
        {
            "pdf": "B.pdf",
            "start_balance": 999.0,
            "end_balance": 1111.0,
            "continuity_start_balance": 999.0,
            "date_min": date(2024, 2, 1),
            "date_max": date(2024, 2, 28),
        },
    ]
    links_bad = compute_statement_continuity(rr_bad)
    assert links_bad and links_bad[0]["status"] in ("Mismatch", "Not checked"), links_bad

    # Money parsing: 0.00 is valid; "£0.00" should not be treated as missing
    rr_zero = [
        {
            "pdf": "Z1.pdf",
            "start_balance": "£0.00",
            "end_balance": "£0.00",
            "date_min": date(2024, 1, 1),
            "date_max": date(2024, 1, 31),
        },
        {
            "pdf": "Z2.pdf",
            "start_balance": "0.00",
            "end_balance": "0.00",
            "date_min": date(2024, 2, 1),
            "date_max": date(2024, 2, 28),
        },
    ]
    links_zero = compute_statement_continuity(rr_zero)
    assert links_zero and links_zero[0]["status"] in ("OK", "Mismatch"), links_zero

    tx = [
        {
            "Date": date(2024, 1, 1),
            "Transaction Type": "CARD",
            "Description": "Tesco",
            "Amount": -10.0,
            "Balance": 90.0,
        },
        {
            "Date": date(2024, 1, 2),
            "Transaction Type": "BACS",
            "Description": "PAY",
            "Amount": 20.0,
            "Balance": 110.0,
        },
    ]
    fp1 = compute_statement_fingerprint(tx)
    fp2 = compute_statement_fingerprint(list(reversed(tx)))
    assert fp1 == fp2 and fp1, (fp1, fp2)

    assert sanitize_filename('A<B>"C"') == "A_B__C_", sanitize_filename('A<B>"C"')

    # Additional test: period extraction for Starling summary line
    # (Only runs when pdfplumber is available; otherwise it's a no-op)
    if _require_pdfplumber(show_error=False) is None:
        assert extract_statement_period_from_pdf("dummy.pdf", bank_hint="starling") == (None, None)

    # reconcile_statement should pick up parser.extract_statement_period when present
    class _FakeParser:
        __name__ = "starling"

        @staticmethod
        def extract_statement_period(_pdf_path: str):
            return (date(2024, 1, 1), date(2024, 1, 31))

        @staticmethod
        def extract_statement_balances(_pdf_path: str):
            return {"start_balance": 100.0, "end_balance": 110.0}

    r = reconcile_statement(_FakeParser, "dummy.pdf", [{"Amount": 10.0}])
    assert r.get("period_start") == date(2024, 1, 1), r
    assert r.get("period_end") == date(2024, 1, 31), r
    assert r.get("status") == "OK", r

    print("Self-tests passed.")
